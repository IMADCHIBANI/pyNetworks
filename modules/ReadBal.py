from openpyxl import Workbook
from openpyxl import load_workbook
from iteration_utilities import duplicates
from iteration_utilities import unique_everseen
from openpyxl.styles import Color, PatternFill, Font, Border
from ..modules.adresse import readAdresse
from ..modules.nommageCable import generateCables
import re

""" readBal est une classe qui permet de génerer un excel contenant les nommages pour les champs des Sitesclients et les cables dans Networks/Netdesigner  """
class readBal:
    def __init__(self,bal_file,Pm,stringSave,synoptique,SROBPI,CodeCableSFR):
        self.bal_file=bal_file
        self.Pm=Pm
        self.SROBPI=SROBPI
        self.stringSave=stringSave
        self.synoptique=synoptique
        self.CodeCableSFR=CodeCableSFR
    """ Fonction pour lire les données et créer l'excel avec les resultats final """
    def readfile(self):
        wb=load_workbook(self.bal_file)
        ws=wb.active
        listofAllPbo=[]
        AllList=[]
        # lire le fichier excel et mettre dans une liste nommée (listeofAllPbo) tout les Pbo du meme Pm qui se trouve dans la cellule 21
        for cell in ws['A2':"V"+str(ws.max_row)]:
            if cell[16].value==self.Pm:
                listofAllPbo.append(cell[21].value)

        #Chaque Pbo peut se répeter dans plusieurs cellules, pour cela la fonction unique_everseen renvoie une liste des Pbo unique.
        listeOfUniquePbo=list(unique_everseen(listofAllPbo))



        for i in range(0,len(listeOfUniquePbo)):
            AllList.append(self.return_IdRA_Adress_of_each_pbo(ws,listeOfUniquePbo[i]))

        self.create_excel(AllList,self.stringSave,ws)
        print("finish")

    """Une fonction qui prend en entrée une liste des Pbo et qui renvoie une liste des idRa + adresses +le nombres de logement qui corresponde à chaque Pbo"""
    def return_IdRA_Adress_of_each_pbo(self,ws,listeOfUniquePbo):
        listofAdresseAndIDRA=[]
        listofIDRA=[]
        listeofAdresse=[]
        listOfDistribution=[]
        listOfLogement=[]

        for cell in ws['A2':"V"+str(ws.max_row)]:
            if cell[16].value==self.Pm:
                if cell[21].value==listeOfUniquePbo:
                    try:
                        cell5=cell[4].value
                        idRa=cell5.split("/")
                        #les idRa se trouve dans la cellule  4 ,mais parfois dans la cellule 5
                        if idRa[0]=="IMB":
                            AdresseAndIDRA=[cell[0].value,cell[5].value,cell[18].value,cell[10].value]

                        else:
                            AdresseAndIDRA=[cell[0].value,cell[4].value,cell[18].value,cell[10].value]
                            #print(AdresseAndIDRA)

                        #Mettre dans une liste de cette forme [ [],[],[],....]  des listes qui contiennent l'adresse + l'idRa + la nature de la distribution et le nombre de logement qui corresponde à chaque PBO.
                        listofAdresseAndIDRA.append(AdresseAndIDRA)
                    except :
                        cell5=cell[5].value
                        idRa=cell5.split("/")
                        #les idRa se trouve dans la cellule  4 ,mais parfois dans la cellule 5
                        if idRa[0]=="IMB":
                            AdresseAndIDRA=[cell[0].value,cell[4].value,cell[18].value,cell[10].value]

                        else:
                            AdresseAndIDRA=[cell[0].value,cell[5].value,cell[18].value,cell[10].value]
                            #print(AdresseAndIDRA)

                        #Mettre dans une liste de cette forme [ [],[],[],....]  des listes qui contiennent l'adresse + l'idRa + la nature de la distribution et le nombre de logement qui corresponde à chaque PBO.
                        listofAdresseAndIDRA.append(AdresseAndIDRA)



        # En fonction du nombre des PBO, on essaye ici de séparer  dans des listes les adresses, les idRa, distribution et les nombre de logement pour chaque PBO
        for i in range(0,len(listofAdresseAndIDRA)):
            listeofAdresse.append(listofAdresseAndIDRA[i][0])
            listofIDRA.append(listofAdresseAndIDRA[i][1])
            listOfDistribution.append(listofAdresseAndIDRA[i][2])
            listOfLogement.append(listofAdresseAndIDRA[i][3])

        #Joindre les Idra pour chaque PBO avec des points vergules ";"
        IdRaString=";".join(map(str, listofIDRA))
        ville=ws["C2"].value
        try:
            #La fonction readFullListe de la classe readAdresse prend en entrée les Adresses de chaque PBO et les regroupes dans une seule Adresse.
             #  finishedlist= Mettre dans une liste final tout les infos concernant un PBO [Nom du PBO,les idRA du PBO,les adresses du PBO,nature de
              # distribution du PBO,la somme des nombres du logement de chaque PBO (générer par la fonction readLogement)]
            finishedlist=[listeOfUniquePbo,IdRaString,readAdresse(self.returnAdress(listeofAdresse)).readFullListe()+" "+str(ws["B2"].value)+" "+str(ville.upper()),listOfDistribution,self.readLogement(listOfLogement)]
            print(finishedlist)
            return finishedlist
        except:
            #En cas d'une erreur retourner une liste vide
            finishedlist=["X",IdRaString," "+" "+" "+" "+" "]
            #print(finishedlist,"erreur")
            return finishedlist

    """Fonction qui prend en entrée une liste de nombre de logement et retourne en sortie la somme de ces logements."""
    def readLogement(self,listOfLogement):
        logement=0
        for i in range(0,len(listOfLogement)):
            logement=listOfLogement[i]+logement
        return logement

    """Fonction qui sépare les numéros de voie des Adresses
       Exemple : [ [25 B RUE IMAD CHIBANI],[26 RUE IMAD CHIBANI] ]--> returnAdress donne le résulat suivant:[ ['25','B','RUE IMAD CHIBANI'],['26','RUE IMAD CHIBANI'] ]"""
    def returnAdress(self,listeofAdresse):
        testcombinedadresse=[]
        for adresse in listeofAdresse:
            if len(adresse.split(maxsplit=2)[1])==1:
                testcombinedadresse.append(adresse.split(maxsplit=2))

            else:
                testcombinedadresse.append(adresse.split(maxsplit=1))

        return testcombinedadresse

    """Création des cellules avec des propriété prédéfinie comme la couleur,le font,la largeur et les valeurs souhaitées. """
    def createcell(self,SiteClient,OrangeFill,font,r,column,Number,width,value1,value2):
        SiteClient.cell(row=1,column=column).value=value1
        SiteClient.cell(row=1,column=column).fill=OrangeFill
        SiteClient.cell(row=1,column=column).font=font
        SiteClient.cell(row=r,column=column).value=value2
        SiteClient.column_dimensions[Number].width = width
    """Fonction pour créer l'excel avec les trois sheets : SiteClient,Cable et l'Intrasite """
    def create_excel(self,AllList,stringSave,ws):
        workbook=Workbook()
        SiteClient=workbook.active
        SiteClient.title="SiteClient"
        Cable=workbook.create_sheet()
        Cable.title="Cable"
        IntraSite=workbook.create_sheet()
        IntraSite.title="IntraSite"

        max=len(AllList)
        #création d'une cellule avec la couleur orange
        OrangeFill = PatternFill(start_color='FF8C00',end_color='FF8C00',fill_type='solid')
        #création d'un font pour la cellule
        font = Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        #Initialisation du paramètre r : création des cellules à partir de la deuxième ligne
        r = 2
        for pbo in AllList:

            self.createcell(SiteClient,OrangeFill,font,r,1,'A',10,"Status","New")
            #------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,2,'B',10,"RESEAU","FTTH")
            #------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,3,'C',15,"CODE_PROJET","NA")
            #-----------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,4,'D',10,"PRECISION","100%")
            #---------------------------------------------------
            #Remplacement du nom du PBO en ajoutant le code du SROBPI
            #Exemple: [PBO-1_AERIEN_B_mM]--> [PBO-SRO-BPI-81065989-001]

            try:
                pbtest=pbo[0]
                Nompbo=pbtest.split("_")
                nbrPbO=Nompbo[0].split("-")
                #Si PBO-X remplacer par PBO-00X
                #si PBO-XX REMPLACER PAR PBO-0XX
                #si PBO-xxx ne rien remplacer
                if len(nbrPbO[1])==1:
                    nbrPbO[1]="00"+nbrPbO[1]
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-"+nbrPbO[1]
                    self.createcell(SiteClient,OrangeFill,font,r,5,'E',30,"USER REFERENCE",nom)
                elif len(nbrPbO[1])==2:
                    nbrPbO[1]="0"+nbrPbO[1]
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-"+nbrPbO[1]
                    self.createcell(SiteClient,OrangeFill,font,r,5,'E',30,"USER REFERENCE",nom)
                else:
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-"+nbrPbO[1]
                    self.createcell(SiteClient,OrangeFill,font,r,5,'E',30,"USER REFERENCE",nom)
            except :
                self.createcell(SiteClient,OrangeFill,font,r,5,'E',30,"USER REFERENCE","??")
            #SiteClient.cell(row=r,column=4).value=pbo[2]
            #--------------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,6,'F',20,"FONCTION_DU_SITE","GP FTTH")
            #--------------------------------------------------------------
            try:
                #Tant que la la cellule de la nature de distribution n'est pas vide :

                if pbo[3][0]!=None:
                    #Si la nature de distribution est Aerien remplacer par "PB Aerien"
                    if pbo[3][0]=="AE_FT":
                        self.createcell(SiteClient,OrangeFill,font,r,7,'G',20,"STRUCTURE_DU_SITE","PB Aerien")
                    #Si la nature de distribution est Souterraine remplacer par "PB Chambre"
                    elif pbo[3][0]=="GC_FT":
                        self.createcell(SiteClient,OrangeFill,font,r,7,'G',20,"STRUCTURE_DU_SITE","PB Chambre")
                    #Si la nature de distribution est Facade (dans ce cas dans le fichie BAL c'est écrit MIXTE) remplacer par "PB Facade"
                    elif pbo[3][0]=="MIXTE":
                        self.createcell(SiteClient,OrangeFill,font,r,7,'G',20,"STRUCTURE_DU_SITE","PB Facade")
                #Sinon remplacer par un point d'interogation pour indiquer que la nature de distribution est inconnue
                else:
                    self.createcell(SiteClient,OrangeFill,font,r,7,'G',20,"STRUCTURE_DU_SITE","?")

            except :pass
            #--------------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,8,'H',20,"PROPRIETAIRE","SFR_FTTH_ZMD_Legacy")
            #---------------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,9,'I',20,"GESTIONNAIRE","SFR_FTTH_ZMD_Legacy")
            #------------------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,10,'J',90,"ADRESSE1",pbo[2])
            #------------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,11,'K',70,"IDADRESSE1",pbo[1])
            #---------------------------------------------------
            try:
                #Parfois la valeur de la celulle du nombre de logement n'est pas un entier donc pour éviter tout cela convertir tout le nombre en format string
                self.createcell(SiteClient,OrangeFill,font,r,12,'L',10,"NOMBRE_LOGEMENT",str(pbo[4]))
            except : pass
            #---------------------------------------------------
            self.createcell(SiteClient,OrangeFill,font,r,13,'M',10,"BAILLEUR","AUTRES")
            #------------------------------------------------------
            #La ligne est remplie donc -> Incrémenter le nombre de ligne pour passer à la ligne suivante
            r += 1


        ville=ws["C7"].value
        self.save_cables(Cable,self.synoptique,ville.upper(),self.CodeCableSFR,OrangeFill,font)
        self.CreateIntraSite(IntraSite,OrangeFill,font,AllList)
        workbook.save(stringSave)
    """Fonction qui à partir du fichier synoptique gènere le nommage de chaque cable entre deux PBO  """
    def save_cables(self,Cable,stringSave,ville,codeSFR,OrangeFill,font):
        #la fonction readfile de la class generateCables prend comme entré le lien du synoptique + la ville (à partie de la cellule C7 du fichier BAL) ainsi que le code du cable SFR (Unique pour chaque NRO) et renvoie une liste de chaque cables
        listOfCables=generateCables(stringSave,ville,codeSFR,self.SROBPI).readfile()
        r=2
        for cable in listOfCables:

            self.createcell(Cable,OrangeFill,font,r,1,'A',80,"REFRENCE",str(cable[0])+" / "+str(cable[1])+" / "+ville)
            #--------------------------------------------
            self.createcell(Cable,OrangeFill,font,r,2,'B',80,"USER REFRENCE","SFR "+str(codeSFR)+" "+cable[2]+" FO -A "+cable[0]+" -O "+cable[1]+" E")
            #--------------------------------------------
            if cable[3]=="SOUTERRAIN":
                self.createcell(Cable,OrangeFill,font,r,3,'C',15,"EMPRISE","GC FT")
            else:
                self.createcell(Cable,OrangeFill,font,r,3,'C',15,"EMPRISE","AERIEN")
            r += 1
    """ Fonction pour Créer le sheet de l'Intrasite """
    def CreateIntraSite(self,IntraSite,OrangeFill,font,allListe):
        r=2
        for list in allListe:
            #print(list[0],list[3][0],list[4])
            pboList=list[0].split("_")
            pbo=pboList[0].split("-")


            self.createcell(IntraSite,OrangeFill,font,r,1,'A',10,"SiteClient"," ")
            self.createcell(IntraSite,OrangeFill,font,r,2,'B',10,"Batiment","HABITATION")
            self.createcell(IntraSite,OrangeFill,font,r,3,'C',10,"Escalier","HABITATION")
            self.createcell(IntraSite,OrangeFill,font,r,4,'D',10,"Etage","HABITATION")
            try:
                #list[4]-->Nombre de logement
                self.createcell(IntraSite,OrangeFill,font,r,5,'E',10,"Nbre de prise",list[4])
            except :pass
            self.createcell(IntraSite,OrangeFill,font,r,6,'F',10,"Code Prise",5100)
            self.createcell(IntraSite,OrangeFill,font,r,7,'G',10,"LOCAL BE","LOCAL BE")
            self.createcell(IntraSite,OrangeFill,font,r,8,'H',30,"Adresse"," ")
            try:
                if len(pbo[1])==1:
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-00"+pbo[1]
                    self.createcell(IntraSite,OrangeFill,font,r,9,'I',30,"Nom de PBO",nom)
                    self.createcell(IntraSite,OrangeFill,font,r,13,'M',30,"Be à relier",nom)
                elif len(pbo[1])==2:
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-0"+pbo[1]
                    self.createcell(IntraSite,OrangeFill,font,r,9,'I',30,"Nom de PBO",nom)
                    self.createcell(IntraSite,OrangeFill,font,r,13,'M',30,"Be à relier",nom)
                else:
                    nom="PBO-SRO-BPI-"+str(self.SROBPI)+"-"+pbo[1]
                    self.createcell(IntraSite,OrangeFill,font,r,9,'I',30,"Nom de PBO",nom)
                    self.createcell(IntraSite,OrangeFill,font,r,13,'M',30,"Be à relier",nom)
                #-----------------------------------
                if list[3][0]=="GC_FT":
                    #Si la distribution est souterraine mettre le code 4824
                    self.createcell(IntraSite,OrangeFill,font,r,10,'J',10,"Code PBO",4824)
                    self.createcell(IntraSite,OrangeFill,font,r,15,'O',10,"Emprise","CHAMBRE")

                elif list[3][0]=="AE_FT":
                    #si elle est aérienne mettre 4820
                    self.createcell(IntraSite,OrangeFill,font,r,10,'J',10,"Code PBO",4820)
                    self.createcell(IntraSite,OrangeFill,font,r,15,'O',10,"Emprise","POTEAU")
                else:
                    #Si elle Facade mettre 4820
                    self.createcell(IntraSite,OrangeFill,font,r,10,'J',10,"Code PBO",4820)
                    self.createcell(IntraSite,OrangeFill,font,r,15,'O',10,"Emprise","FACADE")
            except : pass

            #------------------------------------
            self.createcell(IntraSite,OrangeFill,font,r,11,'K',10,"Cassette","CASSETTE-01")
            self.createcell(IntraSite,OrangeFill,font,r,12,'L',10,"Code Cassette",5600)


            self.createcell(IntraSite,OrangeFill,font,r,14,'N',10,"Type de liaison","Jumper")

            r+=1
"""
Bal="C:/Users/ICHIBANI/Desktop/DOSSIER PBO/ST AUNES/PM17822/Export bal 17822.xlsx"
CodeCableSFR="3410002300"
Pm="PM17822"
stringsave="C:/Users/ICHIBANI/Desktop/DOSSIER PBO/ST AUNES/PM17822/PM17822.xlsx"
synoptique="C:/Users/ICHIBANI/Desktop/DOSSIER PBO/ST AUNES/PM17822/ST AUNES_PM17822_SYNO_DOE.xlsx"
SROBPI="34240822"
v1=readBal(Bal,Pm,stringsave,synoptique,SROBPI,CodeCableSFR).readfile()"""
