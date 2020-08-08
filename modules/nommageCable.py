from openpyxl import Workbook
from openpyxl import load_workbook
import re
class generateCables:
    def __init__(self,doe_file,ville,codeSFR,sroBpi):
        self.doe_file=doe_file
        self.ville=ville
        self.codeSFR=codeSFR
        self.sroBpi=sroBpi
    """Fonction qui permet de génerer les nommages du cables Entre deux PBO à remplire dans la Réference et UserRéference sur Networks/Netdesigner """
    def readfile(self):
        wb=load_workbook(self.doe_file)
        ws=wb["LISTE"]
        listCable=[]
        listCable_with_capacity=[]
        listeofPboAndLogement=[]

        for cableName in ws['A2':"k"+str(ws.max_row)]:
            #if re.search(r'FT',cableName[0].value):
            #FO= Capacité du cable se trouve dans la quatrieme collone du sheet LISTE
            FO=str(cableName[3].value)
            #cable_capacity= un string contenant le nommage du cable et la nature du distribution du cable

            cable=cableName[0].value
            cablesepareted=cable.split("/")
            if len(cablesepareted)==2:
                if re.search(r'PBO',cablesepareted[0]):
                    boiteA=cablesepareted[0].split("-")
                    boiteB=cablesepareted[1].split("-")
                    NommageBoiteA="PBO-SRO-BPI-"+self.sroBpi+"-"+boiteA[4]
                    NommageBoiteB="PBO-SRO-BPI-"+self.sroBpi+"-"+boiteB[4]
                    NommageFinalDuCable=NommageBoiteA+"/ "+NommageBoiteB
                    cable_capacity=NommageFinalDuCable+" "+FO+" "+cableName[9].value
                    cable_table=cable_capacity.split()
                    Name_Lg_Etat=cableName[4].value,cableName[10].value,cableName[9].value

                    for i in range(len(cable_table)):
                        listCable.append(cable_table[i])
                else:
                    #print(cablesepareted)
                    if re.search(r'PBO',cablesepareted[1]):
                        boiteA=cablesepareted[0]
                        boiteB=cablesepareted[1].split("-")
                        NommageBoiteA=boiteA
                        NommageBoiteB="PBO-SRO-BPI-"+self.sroBpi+"-"+boiteB[4]
                        NommageFinalDuCable=NommageBoiteA+"/ "+NommageBoiteB
                        cable_capacity=NommageFinalDuCable+" "+FO+" "+cableName[9].value
                        cable_table=cable_capacity.split()
                        #print(cable_capacity)
                        Name_Lg_Etat=cableName[4].value,cableName[10].value,cableName[9].value
                        for i in range(len(cable_table)):
                            listCable.append(cable_table[i])
                    else:
                        NommageFinalDuCable=cablesepareted[0]+"/"+cablesepareted[1]
                        cable_capacity=NommageFinalDuCable+" "+FO+" "+cableName[9].value
                        cable_table=cable_capacity.split()
                        #print(cable_capacity)
                        Name_Lg_Etat=cableName[4].value,cableName[10].value,cableName[9].value
                        for i in range(len(cable_table)):
                            listCable.append(cable_table[i])

            else:
                boite=cable.split(" ")
                boiteA=boite[0]
                boiteB=boite[2]
                numeroBoiteB=boiteB.split("-")
                if re.search(r'PBO',boiteB):
                    NommageBoiteA=boiteA
                    NommageBoiteB="PBO-SRO-BPI-"+self.sroBpi+"-"+numeroBoiteB[4]
                    NommageFinalDuCable=NommageBoiteA+" / "+NommageBoiteB
                    cable_capacity=NommageFinalDuCable+" "+FO+" "+cableName[9].value
                    cable_table=cable_capacity.split()
                    #print(cable_capacity)
                    Name_Lg_Etat=cableName[4].value,cableName[10].value,cableName[9].value
                    for i in range(len(cable_table)):
                        listCable.append(cable_table[i])
                else:
                    cable_capacity=cable+" "+FO+" "+cableName[9].value
                    cable_table=cable_capacity.split()
                    Name_Lg_Etat=cableName[4].value,cableName[10].value,cableName[9].value
                    for i in range(len(cable_table)):
                        listCable.append(cable_table[i])

        listofeachrow=[]

        for i in range(0,int(len(listCable))):

            try:
                listerow=[]
                listerow.append(listCable[0+5*i])
                listerow.append(listCable[2+5*i])
                listerow.append(listCable[3+5*i])
                listerow.append(listCable[4+5*i])

                listofeachrow.append(listerow)
                print(listerow)
            except IndexError:
                break
        #print(listofeachrow)
        return listofeachrow

#v1=generateCables("C:/Users/ICHIBANI/Desktop/DOSSIER PBO/Mantaud/PM001/MONTAUD_PM01_SYNO_DOE.xlsx","MONTAUD","3410002299","imad").readfile()
